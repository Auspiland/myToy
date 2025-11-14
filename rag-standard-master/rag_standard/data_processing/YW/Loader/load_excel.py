import os, re, openpyxl, xlrd, olefile, sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook, chartsheet
from fractions import Fraction
from datetime import datetime

# 한글이나 중국어 등 아시아권 폰트가 포함된 엑셀 파일의 경우, 폰트의 ptichFamily 값이 Openpyxl의 기본 최대값(52)을 초과하는 경우 있음 - 그래서 제한 늘려서 오류 방지
openpyxl.drawing.text.Font.pitchFamily.max = 1000
print("Modified Openpyxl")

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../..')) # 상대경로로 root 설정 (Rag_Standard_v2)
sys.path.append(BASE_DIR)
from rag_standard.utils import setup_logger

global log
log = setup_logger.setup_logger('LoadEXCEL', 'load_excel')

class LoadExcel:
    """
    load_excel만 호출하면 됨
    - 필요한 경우 convert_xls_to_xlsx로 변환해주거나, xls_meta로 xls 문서의 날짜를 추출
    - is_hidden으로 숨김 처리된 행이나 열 처리
    - format_value로 셀 값 포맷팅 <- convert_to_mixed_fraction: 대분수로 변환하는 함수
    """
    def __init__(self):
        pass

    def load_excel(self, file_path, mid_path):
        """
        엑셀 파일(.xls, .xlsx, .xlsm)을 로드하고 내용을 추출하는 메소드 >> Openpyxl(기본), xlrd(xls를 변환할 때만) 사용
        
        Args:
            file_path: 엑셀 파일 경로
            mid_path: 변환된 파일을 저장할 출력 폴더 경로
            
        Returns:
            tuple: (파일명, 확장자, 수정일, 생성일, 텍스트만 추출한 딕셔너리, 텍스트와 테이블을 모두 추출한 딕셔너리, 전체 텍스트를 저장한 문자열)
            - 실패 시 False 반환

        딕셔너리 형식:
            {
                '시트명': 시트 내용,
                '시트명': 시트 내용,
                ...
            }
        """
        file_name_without_ext, extension = os.path.basename(file_path).rsplit('.', 1) # 파일명과 확장자 분리

        try:
            # xls 파일인 경우 xlsx로 변환 (오래된 형식의 엑셀 파일 처리)
            if extension.lower() == 'xls':
                # 변환된 파일을 저장할 디렉토리 생성
                new_dir = f"{mid_path}/xls_convert"
                os.makedirs(new_dir, exist_ok=True)
                to_xlsx_file = os.path.join(new_dir, f'{file_name_without_ext}.xlsx')
                # 변환 실패시 False 반환
                if not self.convert_xls_to_xlsx(file_path, to_xlsx_file): 
                    log.info(f"Failed to convert XLS file: {file_path}")
                    return False
                # XLS 파일의 메타데이터(수정일, 생성일) 추출
                modified_date, created_date = self.xls_meta(file_path)
                file_to_load = to_xlsx_file
                
                # 변환된 XLSX 파일 로드 (수식 결과값만 로드)
                workbook = load_workbook(file_to_load, data_only=True)
                
            else:
                # xlsx, xlsm 파일 처리 (최신 형식의 엑셀 파일)
                file_to_load = file_path
                workbook = load_workbook(file_to_load, data_only=True)
                try:
                    # 워크북 메타데이터에서 날짜 정보 추출
                    if workbook.properties.modified:
                        modified_date = workbook.properties.modified.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        # 메타데이터에 수정일이 없는 경우 파일 시스템의 수정일 사용
                        modified_date = os.path.getmtime(file_path).strftime("%Y-%m-%d %H:%M:%S")
                        log.info(f"Modified date not found in metadata for {file_path}, using file system date")

                    if workbook.properties.created:
                        created_date = workbook.properties.created.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        # 메타데이터에 생성일이 없는 경우 파일 시스템의 생성일 사용
                        created_date = os.path.getctime(file_path).strftime("%Y-%m-%d %H:%M:%S")
                        log.info(f"Created date not found in metadata for {file_path}, using file system date")
                except (AttributeError, TypeError) as e:
                    # 메타데이터 접근 실패시 파일 시스템의 날짜 정보 사용
                    log.warning(f"Error accessing workbook properties for {file_path}: {e}")
                    modified_date = os.path.getmtime(file_path).strftime("%Y-%m-%d %H:%M:%S")
                    created_date = os.path.getctime(file_path).strftime("%Y-%m-%d %H:%M:%S")

            # 시트별 데이터를 저장할 딕셔너리 초기화
            text_only_sheets = {} # 텍스트만 추출한 내용
            text_and_table_sheets = {} # 텍스트와 표를 추출한 내용

            # 각 시트 처리
            for sheet_name in workbook.sheetnames:
                # 보이는 시트만 처리 (숨김 시트 제외)
                if workbook[sheet_name].sheet_state == 'visible':
                    sht = workbook[sheet_name]

                    # 차트 시트는 건너뛰기 - 엑셀 시트가 차트 시트인지 확인
                    if isinstance(sht, chartsheet.Chartsheet):
                        log.info(f"Skipping chart sheet: {sheet_name}")
                        continue
                        
                    # 데이터 읽기 (숨김 처리된 행과 열 제외)
                    data = []
                    hidden_rows = [row for row in range(1, sht.max_row + 1) if self.is_hidden(sht, row, 'row')]

                    # 모든 셀 데이터 읽기
                    for row_idx, row in enumerate(sht.iter_rows(), start=1):
                        if row_idx not in hidden_rows: # 숨김 처리된 행 제외
                            row_data = []
                            for col_idx, cell in enumerate(row, start=1):
                                    if not self.is_hidden(sht, col_idx, 'column'): # 숨김 처리된 열 제외
                                        value = self.format_value(cell.value, cell.number_format) # 셀 값 포맷팅
                                        row_data.append(value)
                            if any(row_data): # 비어있지 않은 행만 추가
                                data.append(row_data)
        
                    sheet = pd.DataFrame(data) # 데이터프레임으로 변환
        
                    # NaN 처리 및 문자열 변환 (NaN을 None으로 변환 후 모든 셀을 문자열로 변환하고 전처리)
                    sheet = sheet.where(pd.notna(sheet), None) # NaN을 None으로 변환
                    sheet = sheet.apply(lambda col: col.astype(str))  # 모든 셀을 문자열로 변환하고 전처리
                    sheet_reset = sheet.reset_index(drop=True) # 헤더를 데이터 행으로 포함
        
                    # 빈 값 처리
                    # pd.set_option('future.no_silent_downcasting', True) # 주석처리 (0317)
                    sheet_reset.replace(to_replace=['None', 'nan', '', ' '], value=np.nan, inplace=True) # 'None', 'nan', ' ', ''을 NaN으로 변환
                
                    # 빈 열과 행 제거
                    sheet_reset.dropna(axis=0, how='all', inplace=True)  # 모든 값이 NaN인 행 제거
                    sheet_reset.dropna(axis=1, how='all', inplace=True)  # 모든 값이 NaN인 열 제거

                    # Unnamed 열 처리 - 빈 문자열로 변경
                    sheet_reset = sheet_reset.apply(lambda col: col.map(lambda x: None if isinstance(x, str) and x.startswith('Unnamed:') else x))
                    sheet_value = sheet_reset.dropna(how='all', axis='columns').dropna(how='all')

                    # 비어있지 않은 시트 처리
                    if not sheet_value.empty:
                        # 헤더 설정
                        new_header = sheet_value.iloc[0] # 첫 번째 행을 새 헤더로 설정
                        sheet_set = sheet_value[1:] # 첫 번째 행 제거
                        sheet_set.columns = new_header

                        # 텍스트 추출 (모든 셀 값을 결합하여 단순 텍스트로 반환)
                        sheet_text = sheet_set.fillna('')
                        text_only_content = "\n".join([" ".join(map(str, row)) for row in sheet_text.values])
                        text_only_sheets[sheet_name] = re.sub(r' {2,}', ' ', text_only_content).strip()

                        # 마크다운 테이블 생성
                        table_markdown = sheet_set.to_markdown(index=False)
                        markdown_final = re.sub(r'\|:?-+:?', '|---', re.sub(r'\s{2,}', ' ', re.sub(r'nan', ' ', re.sub(r'None', ' ', table_markdown))))                        
                        text_and_table_sheets[sheet_name] = markdown_final
                    else: # 빈 시트인 경우
                        text_only_sheets[sheet_name] = ""
                        text_and_table_sheets[sheet_name] = ""

                text_and_table_result = "\n\n".join(text_and_table_sheets.values())
                
            log.info(f"Completed loading excel file: {os.path.basename(file_path)}")
                    
        except Exception as e:
            log.info(f"Cannot load EXCEL file: {file_path} due to {e}")
            return False
            
        return (file_name_without_ext, extension.lower(), modified_date, created_date, text_only_sheets, text_and_table_sheets, text_and_table_result)

    def convert_xls_to_xlsx(self, xls_file, xlsx_file):
        """
        xls 형태의 파일을 xlsx 형태의 파일로 변환
        xls 형태의 파일을 로드하는 xlrd를 사용하면, 로더에서 셀에 기입된 날짜 형태의 정보 인식 못함 -> 그래서 xlsx 형태로 변환 후 로드 -> 날짜 제대로 기입된 대로 읽힘

        Args:
            xls_file: 변환할 xls 파일 경로
            xlsx_file: 저장할 xlsx 파일 경로
        
        Returns:
            bool: 변환 성공 여부
        """
        # 파일명과 확장자 분리
        file_name_without_ext, extension = os.path.basename(xls_file).rsplit('.', 1)

        try:
            # xls 파일 로드
            xls_book = xlrd.open_workbook(xls_file, on_demand=True)
            # 새로운 xlsx 워크북 생성
            workbook = Workbook()
        
            first_sheet = True
            for xls_sheet in xls_book.sheets():
                # 숨김처리된 시트는 제외하고 처리
                if not xls_sheet.visibility:
                    if first_sheet:
                        # 첫 번째 시트는 기본 시트를 사용
                        sheet = workbook.active
                        sheet.title = xls_sheet.name
                        first_sheet = False
                    else:
                        # 새로운 시트 생성
                        sheet = workbook.create_sheet(title=xls_sheet.name)

                    # 모든 셀을 순회하면서 데이터 복사
                    for row in range(xls_sheet.nrows):
                        for col in range(xls_sheet.ncols):
                            cell_value = xls_sheet.cell_value(row, col)
                            cell_type = xls_sheet.cell_type(row, col)
        
                            # 날짜 타입의 셀은 년-월-일 형식의 문자열로 변환
                            if cell_type == xlrd.XL_CELL_DATE:
                                log.info(f"xls 변환 중 ... sheet_name 출력 : {sheet.title}")
                                log.info(f"xls 변환 중 ... cell_value 출력 : {cell_value}")
                                # date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                                data_tuple1 = xlrd.xldate_as_datetime(cell_value, xls_book.datemode)
                                # cell_value = datetime(*date_tuple)
                                cell_value = data_tuple1.strftime('%Y-%m-%d')  # 년-월-일 형식으로 변환
                                log.info(f"xls 변환 중 ... 변경된 cell_value 출력 : {cell_value}")
        
                            # openpyxl의 column 키워드를 사용하여 셀에 값 입력
                            sheet.cell(row=row + 1, column=col + 1, value=cell_value)

            # 새로운 xlsx 파일로 저장
            workbook.save(xlsx_file)
            return True
        
        except Exception as e:
            log.info(f"Error converting {xls_file} to {xlsx_file}: {e}")
            return False
        
    def xls_meta(self, file):
        """
        xls 파일의 생성일자와 수정일자를 추출하는 함수
        
        Args:
            file: xls 파일 경로
        
        Returns:
            tuple: (수정일자, 생성일자) - Unix timestamp 형식
        """
        try:
            ole = olefile.OleFileIO(file) # OLE 파일 형식으로 열기

            # 메타데이터가 존재하는 경우
            if ole.exists("\x05SummaryInformation"):
                meta = ole.getproperties("\x05SummaryInformation")
                # Windows 시스템 시간(1601년 기준)을 Unix 시간(1970년 기준)으로 변환
                cdate = meta[12] - 11644473612 # 생성일자
                mdate = meta[13] - 11644473612 # 수정일자
            else:
                # 메타데이터가 없는 경우 파일 시스템의 시간 정보 사용
                cdate = os.path.getctime(file)
                mdate = os.path.getmtime(file)

        except Exception as e:
            # 오류 발생 시 파일 시스템의 시간 정보 사용
            cdate = os.path.getctime(file)
            mdate = os.path.getmtime(file)
            log.info(f"Error processing file {file}: {e}")

        return mdate, cdate
    
    def is_hidden(self, sheet, index, dimension_type='row'):
        """
        엑셀 시트의 숨김 처리된 행 또는 열을 찾는 함수
        
        Args:
            sheet: 검사할 엑셀 시트
            index: 검사할 행 또는 열의 인덱스
            dimension_type: 검사 대상('row' 또는 'column') (default: 'row')
        
        Returns:
            bool: 숨김 처리 여부
        """
        if dimension_type == 'row':
            # 행이 존재하면 해당 행의 숨김 여부 반환, 없으면 False 반환
            return sheet.row_dimensions[index].hidden if index in sheet.row_dimensions else False
        elif dimension_type == 'column':
            # 열 인덱스를 문자로 변환 (예: 1 -> 'A')
            col_letter = openpyxl.utils.get_column_letter(index)
            # 열이 존재하면 해당 열의 숨김 여부 반환, 없으면 False 반환
            return sheet.column_dimensions[col_letter].hidden if col_letter in sheet.column_dimensions else False
        
    def format_value(self, value, number_format=None):
        """
        엑셀 셀의 값을 서식에 맞게 처리하는 함수
        
        Args:
            value: 처리할 셀 값
            number_format: 엑셀에 설정된 셀 서식
        
        Returns:
            포맷팅된 값
        """
        # 나눗셈 에러(#DIV/0!) 처리 - 빈 문자열로 변환
        if isinstance(value, str) and value == '#DIV/0!':
            return ''

        # 문자열의 줄바꿈을 공백으로 변환
        if isinstance(value, str):
            return value.replace('\n', ' ')

        # 퍼센트 서식 처리 (예: 0.1 -> "10%")
        if isinstance(value, (int, float)) and number_format and '0%' in number_format:
            return f"{round(value * 100, 1)}%"
        
        # 분수 서식 처리 (예: 1.75 -> "1 3/4")
        elif number_format and (number_format.startswith('# ?/?') or number_format.startswith('0 ?/?')):
            # 소수 값을 대분수로 변환
            return self.convert_to_mixed_fraction(value)
        
        # 날짜 서식 처리
        elif isinstance(value, datetime):
            if number_format and "d" in number_format and "mmm" in number_format:
                return value.strftime("%d-%b") # 일-월 형식 (예: 4-Nov)
            elif number_format and "yyyy" in number_format:
                return value.strftime("%Y-%m-%d") # 년-월-일 형식
            else:
                return value.strftime("%Y-%m-%d") # 기본 날짜 형식 (년-월-일)
        
        # 숫자 값 처리 (소수점 반올림)
        elif isinstance(value, (int, float)):
            # 지정된 소수점 자릿수로 반올림
            if number_format in ['0.0', '#,##0.0']:
                return round(value, 1) # 소수점 1자리
            elif number_format in ['0.00', '#,##0.00']:
                return round(value, 2) # 소수점 2자리
            elif number_format in ['0.000', '#,##0.000']:
                return round(value, 3) # 소수점 3자리
            elif number_format == 'General':
                # 일반 서식에서의 대분수 변환 검토
                if isinstance(value, float) and not value.is_integer():
                    integer_part = int(value)
                    fractional_part = value - integer_part

                    # 소수점 자릿수가 6자리 미만인 경우만 대분수 변환 고려
                    decimal_str = str(fractional_part)
                    decimal_places = len(decimal_str.split('.')[1]) if '.' in decimal_str else 0
                    if decimal_places < 6:
                        fraction = Fraction(fractional_part).limit_denominator()
        
                        # 분모가 8, 16, 32인 간단한 분수만 대분수로 변환
                        if fraction.denominator in [8, 16, 32]:
                            return f"{integer_part} {fraction}"
                        
                # 소수점 자릿수에 따른 반올림
                decimal_str = str(value)
                if '.' in decimal_str:
                    decimal_places = len(decimal_str.split('.')[1])  # 소수점 이하 자릿수 계산
                    if decimal_places == 1:
                        return round(value, 1)
                    elif decimal_places == 2:
                        return round(value, 2)
                return round(value, 1)  # 기본 소수점 1자리
            else:
                return round(value, 1)  # 기본 소수점 1자리
        
        # 그 외 모든 값은 원본 그대로 반환
        return value
    
    def convert_to_mixed_fraction(self, value):
        """
        숫자를 대분수 형태로 변환하는 함수
        (예: 1.75 -> "1 3/4", 0.75 -> "3/4", 2 -> "2")
        
        Args:
            value: 변환할 숫자 또는 문자열
        
        Returns:
            str: 대분수 형태의 문자열 또는 원본 값(변환 실패시)
        """
        try:
            # 값이 숫자나 문자열 형태인지 확인하고 입력값을 분수(Fraction)로 변환 - 분모는 최대 100까지 허용
            if isinstance(value, (int, float, str)):
                fraction_value = Fraction(value).limit_denominator(100)
            else:
                raise ValueError("Input must be a string or a number.")

            # 대분수 형태로 변환 (정수 부분과 분수 부분 분리)
            whole_number = fraction_value.numerator // fraction_value.denominator
            fraction_part = Fraction(fraction_value.numerator % fraction_value.denominator, fraction_value.denominator)

            # 결과 반환 (대분수, 정수, 또는 순수 분수 형태)
            if whole_number != 0 and fraction_part.numerator != 0:
                return f"{whole_number} {fraction_part}"  # 대분수 형태 (예: 1 3/4)
            elif whole_number != 0:
                return f"{whole_number}"  # 정수 형태 (예: 1)
            else:
                return f"{fraction_part}"  # 순수 분수 형태 (예: 3/4)
                
        except ValueError as e:
            # 변환 실패 시 오류 로그를 남기고 원본 값 (원래의 입력값) 반환
            log.error(f"Error converting {value} to mixed fraction: {e} - will just use the input value")
            return value